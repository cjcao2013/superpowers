#!/usr/bin/env python3
"""
TAP Migration Analyzer
Auto-discovers Robot Framework + Excel project structure and generates
TAP requirements doc + converted JSON output.

Usage:
  python analyze.py /path/to/project               # scan only
  python analyze.py /path/to/project --convert-cases
  python analyze.py /path/to/project --convert-data
"""

import os, sys, json, re, argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SENSITIVE_PATTERNS = {"password", "passwd", "token", "secret", "key", "credential", "api_key"}


def is_sensitive(col_name: str) -> bool:
    if not col_name:
        return False
    lower = col_name.lower().replace(" ", "_")
    return any(p in lower for p in SENSITIVE_PATTERNS)


def scan_excel(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(max_row=5, values_only=True))
        if not rows:
            continue
        headers = [str(h) if h is not None else "" for h in rows[0]]
        row_count = ws.max_row - 1 if ws.max_row else 0
        sheets[name] = {
            "headers": headers,
            "row_count": row_count,
            "sensitive_columns": [h for h in headers if is_sensitive(h)],
        }
    wb.close()
    return sheets


def detect_active_column(headers: list[str]) -> tuple[str | None, str | None]:
    """Guess which column flags a test case as active."""
    candidates = ["ExecutorFlag", "Active", "Run", "Execute", "Enabled", "Status", "Flag"]
    for c in candidates:
        for h in headers:
            if h and c.lower() in h.lower():
                return h, "Yes"
    return None, None


def detect_case_id_column(headers: list[str]) -> str | None:
    candidates = ["TC_ID", "TestCase_ID", "Case_ID", "ID", "Test_ID", "TCID"]
    for c in candidates:
        for h in headers:
            if h and c.lower() == h.lower():
                return h
    for h in headers:
        if h and "id" in h.lower():
            return h
    return headers[0] if headers else None


def infer_environment(filename: str) -> str | None:
    """Extract environment name from filename like TestDataFile_STG.xlsx"""
    m = re.search(r'[_-]([A-Z]{2,6})\.(xlsx?|xls)$', filename, re.IGNORECASE)
    return m.group(1).upper() if m else None


def find_rf_run_command(project_root: Path) -> str | None:
    for pattern in ["*.yml", "*.yaml", "Jenkinsfile", "*.sh"]:
        for f in project_root.rglob(pattern):
            try:
                text = f.read_text(errors="ignore")
                m = re.search(r'(robot\s+.*?MainExecutor\S*\.robot[^\n]*)', text)
                if m:
                    return m.group(1).strip()
            except Exception:
                pass
    return None


def detect_rf_pattern(project_root: Path) -> dict:
    """
    Detect if this is Pattern A (standard RF) or Pattern B (single executor + Excel registry).
    Pattern B: only 1 RF test case exists, real cases are in Excel, uses dynamic dispatch.
    """
    robot_files = list(project_root.rglob("*.robot"))
    robot_files = [f for f in robot_files if ".git" not in str(f)]

    total_test_cases = 0
    has_dynamic_dispatch = False
    executor_files = []

    for rf_file in robot_files:
        try:
            text = rf_file.read_text(errors="ignore")
            # Count test cases (lines that look like test case definitions)
            in_test_section = False
            for line in text.splitlines():
                if "*** Test Cases ***" in line:
                    in_test_section = True
                    continue
                if line.startswith("***") and in_test_section:
                    in_test_section = False
                if in_test_section and line and not line.startswith(" ") and not line.startswith("\t") and line.strip():
                    total_test_cases += 1
            # Check for dynamic dispatch: "Run Keyword" followed by anything containing ${...}
            if re.search(r'Run Keyword\s+[^\n]*\$\{', text):
                has_dynamic_dispatch = True
                executor_files.append(str(rf_file.relative_to(project_root)))
        except Exception:
            pass

    pattern = "B" if (total_test_cases <= 2 and has_dynamic_dispatch) else "A"
    return {
        "pattern": pattern,
        "total_rf_test_cases": total_test_cases,
        "has_dynamic_dispatch": has_dynamic_dispatch,
        "executor_files": executor_files,
        "warning": (
            "⚠️  PATTERN B DETECTED: Only 1 RF test case (loop executor). "
            "Real test cases are in Excel registry. "
            "TAP must NOT import RF test cases — import from Excel registry instead."
        ) if pattern == "B" else None,
    }


def analyze(project_root: Path) -> dict:
    result = {
        "project_root": str(project_root),
        "excel_files": {},
        "robot_files": [],
        "binary_fixtures": [],
        "environments": [],
        "rf_run_command": None,
        "test_case_registry": None,
        "test_data_files": {},
        "rf_pattern": None,
    }

    # Find Excel files
    for f in project_root.rglob("*.xlsx"):
        if any(p in str(f) for p in [".git", "__pycache__", "tap-migration"]):
            continue
        rel = str(f.relative_to(project_root))
        sheets = scan_excel(f)
        env = infer_environment(f.name)
        result["excel_files"][rel] = {"sheets": sheets, "environment": env}
        if env:
            if env not in result["environments"]:
                result["environments"].append(env)
            result["test_data_files"][env] = rel

    # Guess which Excel is the test case registry
    for path, info in result["excel_files"].items():
        if info["environment"]:
            continue
        if any(kw in path.lower() for kw in ["testcase", "test_case", "cases", "registry"]):
            result["test_case_registry"] = path
            break
    if not result["test_case_registry"] and result["excel_files"]:
        non_env = [p for p, i in result["excel_files"].items() if not i["environment"]]
        if non_env:
            result["test_case_registry"] = non_env[0]

    # Find Robot Framework files
    for f in project_root.rglob("*.robot"):
        result["robot_files"].append(str(f.relative_to(project_root)))

    # Find binary fixtures
    for ext in ["*.jpg", "*.jpeg", "*.png", "*.pdf", "*.gif"]:
        for f in project_root.rglob(ext):
            if ".git" not in str(f):
                result["binary_fixtures"].append(str(f.relative_to(project_root)))

    # Find RF run command
    result["rf_run_command"] = find_rf_run_command(project_root)

    # Detect RF execution pattern
    result["rf_pattern"] = detect_rf_pattern(project_root)

    return result


def print_report(data: dict):
    print("\n" + "="*60)
    print("TAP MIGRATION SCAN REPORT")
    print("="*60)
    print(f"Project: {data['project_root']}\n")

    pat = data.get("rf_pattern", {})
    if pat:
        print(f"RF EXECUTION PATTERN: Pattern {pat['pattern']}")
        print(f"  RF test cases found: {pat['total_rf_test_cases']}")
        print(f"  Dynamic dispatch (Run Keyword ${{...}}): {pat['has_dynamic_dispatch']}")
        if pat.get("warning"):
            print(f"  {pat['warning']}")
        print()

    if data["test_case_registry"]:
        reg = data["test_case_registry"]
        sheets = data["excel_files"][reg]["sheets"]
        print(f"TEST CASE REGISTRY: {reg}")
        for sheet_name, info in sheets.items():
            headers = info["headers"]
            id_col = detect_case_id_column(headers)
            active_col, active_val = detect_active_column(headers)
            print(f"  Sheet: {sheet_name}")
            print(f"    Columns ({len(headers)}): {', '.join(h for h in headers if h)}")
            print(f"    Case ID column: {id_col}")
            print(f"    Active flag: {active_col} = '{active_val}'")
            print(f"    Rows: ~{info['row_count']}")
            if info["sensitive_columns"]:
                print(f"    ⚠️  Sensitive: {info['sensitive_columns']}")

    print(f"\nENVIRONMENTS: {data['environments'] or ['(none detected)']}")

    print(f"\nTEST DATA FILES:")
    for env, path in data["test_data_files"].items():
        sheets = data["excel_files"][path]["sheets"]
        total_rows = sum(s["row_count"] for s in sheets.values())
        all_sensitive = [c for s in sheets.values() for c in s["sensitive_columns"]]
        print(f"  [{env}] {path}")
        print(f"    Sheets: {list(sheets.keys())}")
        print(f"    Total rows: ~{total_rows}")
        if all_sensitive:
            print(f"    ⚠️  Sensitive columns: {list(set(all_sensitive))}")

    if data["binary_fixtures"]:
        print(f"\nBINARY FIXTURES ({len(data['binary_fixtures'])}):")
        for f in data["binary_fixtures"]:
            print(f"  {f}")

    if data["rf_run_command"]:
        print(f"\nRF RUN COMMAND:\n  {data['rf_run_command']}")
    else:
        print(f"\nRF RUN COMMAND: (not found in CI config — check manually)")

    print("\n" + "="*60)


def convert_cases(data: dict, project_root: Path):
    reg = data["test_case_registry"]
    if not reg:
        print("ERROR: No test case registry found", file=sys.stderr)
        return

    registry_path = project_root / reg
    wb = openpyxl.load_workbook(registry_path, data_only=True)
    sheet_name = list(wb.sheetnames)[0]
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    id_col = detect_case_id_column([h for h in headers if h])
    active_col, active_val = detect_active_column([h for h in headers if h])

    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(headers, row))
        if r.get(id_col) is None:
            continue
        if active_col and str(r.get(active_col, "")).strip().lower() != (active_val or "yes").lower():
            continue
        case = {k: (f"${{{k.upper()}}}" if is_sensitive(k) else v)
                for k, v in r.items() if k}
        cases.append(case)

    out_dir = project_root / "tap-migration"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "testcases.json"
    out_path.write_text(json.dumps(cases, ensure_ascii=False, indent=2, default=str))
    print(f"✓ {len(cases)} active cases → {out_path}")


def convert_data(data: dict, project_root: Path):
    out_dir = project_root / "tap-migration"
    out_dir.mkdir(exist_ok=True)

    for env, rel_path in data["test_data_files"].items():
        wb = openpyxl.load_workbook(project_root / rel_path, data_only=True)
        result = {"environment": env, "sheets": {}}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                r = {h: (f"${{{h.upper()}}}" if is_sensitive(h) else v)
                     for h, v in zip(headers, row) if h}
                rows.append(r)
            result["sheets"][sheet_name] = rows
        wb.close()

        out_path = out_dir / f"testdata_{env}.json"
        out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        total = sum(len(v) for v in result["sheets"].values())
        print(f"✓ [{env}] {len(result['sheets'])} sheets, {total} rows → {out_path}")


def main():
    parser = argparse.ArgumentParser(description="TAP Migration Analyzer")
    parser.add_argument("project_dir", help="Path to the project root")
    parser.add_argument("--convert-cases", action="store_true", help="Convert test cases to JSON")
    parser.add_argument("--convert-data", action="store_true", help="Convert test data to JSON")
    args = parser.parse_args()

    project_root = Path(args.project_dir).resolve()
    if not project_root.exists():
        print(f"ERROR: {project_root} does not exist", file=sys.stderr)
        sys.exit(1)

    data = analyze(project_root)

    if args.convert_cases:
        convert_cases(data, project_root)
    elif args.convert_data:
        convert_data(data, project_root)
    else:
        print_report(data)
        print(json.dumps(data, indent=2))


if __name__ == "__main__":
    main()
